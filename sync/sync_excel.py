"""
sync_excel.py — Sales Activity Console ETL  v3
===============================================
Sources
  Team BP   : SharePoint  logman365.sharepoint.com/sites/BPTeam
              File: Shared Documents/General/TEAM BP.xlsx  (sheet "2026")
  Team MAI  : SharePoint  logman365.sharepoint.com/sites/Salesteam
              File: Shared Documents/General/Sales Log.xlsx
  Team OVS  : Local OneDrive sync (auto-scans ALL monthly sub-folders)
              Sales CRM/TEAM OVERSEA/…/<Month Folder>/<file>.xlsx

SharePoint authentication
  On the FIRST run the script prints a short code and a URL.  Open
  https://microsoft.com/devicelogin, enter the code, and sign in with
  your Microsoft 365 account.  Tokens are cached in sync/.cache so every
  later run is fully automatic (token renews silently for ~90 days).

Run
  cd C:\\Users\\SALES_48\\sales-activity-console
  python sync\\sync_excel.py           # normal run
  python sync\\sync_excel.py --fallback # only runs if last sync failed

Schedule (Task Scheduler)
  Primary  : Tuesday  15:00  → run_sync.bat
  Fallback : Wednesday 10:00 → run_sync_fallback.bat  (skips if already OK)
  Set up with: powershell -ExecutionPolicy Bypass -File sync\\setup_scheduler.ps1
"""

import os
import re
import sys
import json
import math
import subprocess
from datetime import datetime, timedelta

import openpyxl

# =====================================================================
# CONFIGURATION
# =====================================================================

ONEDRIVE_BASE = r"C:\Users\SALES_48\OneDrive - Logman International Co Ltd"
SALES_CRM     = os.path.join(ONEDRIVE_BASE, "Sales CRM")

# SharePoint file URLs (parsed from the .url shortcuts in Sales CRM)
BP_SHAREPOINT_URL  = ("https://logman365.sharepoint.com/sites/BPTeam"
                      "/Shared Documents/General/TEAM BP.xlsx")
MAI_SHAREPOINT_URL = ("https://logman365.sharepoint.com/sites/Salesteam"
                      "/Shared Documents/General/Sales Log.xlsx")

# Local paths — overridable via environment variables (for CI / GitHub Actions)
_SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
_REPO_ROOT  = os.path.dirname(_SCRIPT_DIR)
CACHE_DIR   = os.path.join(_SCRIPT_DIR, ".cache")

BP_PATH  = os.environ.get("BP_PATH",  os.path.join(CACHE_DIR, "TEAM_BP.xlsx"))
MAI_PATH = os.environ.get("MAI_PATH", os.path.join(CACHE_DIR, "Sales_Log.xlsx"))
OVS_BASE = os.environ.get("OVS_BASE", os.path.join(
    SALES_CRM, "TEAM OVERSEA",
    "CHARLIE SK's files - Overseas Sales Pipeline 2026_Shared File",
))
OUTPUT_PATH = os.environ.get(
    "OUTPUT_PATH", os.path.join(_REPO_ROOT, "data.json")
)
MARKET_PATH = os.environ.get(
    "MARKET_PATH", os.path.join(_SCRIPT_DIR, "market_intelligence.xlsx")
)

AUTO_GIT_PUSH = True
TODAY         = datetime.today().date()

# Status file — tracks whether last sync succeeded (used by --fallback mode)
STATUS_FILE = os.path.join(CACHE_DIR, ".last_run_status")

# =====================================================================
# MICROSOFT GRAPH / MSAL SETTINGS
# =====================================================================
# We use Microsoft's own "Azure CLI" public-client app (client ID below).
# This app is pre-consented in virtually every Microsoft 365 tenant and
# requires no Azure App Registration on your part.
#
# If your IT admin has restricted third-party app sign-ins you will see
# an AADSTS error on first run — in that case create a free App
# Registration in your own tenant and paste its client ID here.

MSAL_TENANT    = "logman365.onmicrosoft.com"
MSAL_CLIENT_ID = "e838e2bf-181f-48ca-a6c1-eb40ae5d6c91"   # Sales Activity Sync app
MSAL_SCOPES    = ["https://graph.microsoft.com/Files.Read.All"]
TOKEN_CACHE    = os.path.join(CACHE_DIR, ".token_cache.bin")

# =====================================================================
# STATIC REFERENCE DATA
# =====================================================================

STAGES = [
    {"id": "lead",      "label": "Lead",           "thai": "ลูกค้าใหม่",      "color": "#94a3b8"},
    {"id": "contact",   "label": "First Contact",  "thai": "ติดต่อครั้งแรก",   "color": "#60a5fa"},
    {"id": "visit",     "label": "Site Visit",     "thai": "เข้าพบลูกค้า",     "color": "#22d3ee"},
    {"id": "quote",     "label": "Quotation Sent", "thai": "ส่งใบเสนอราคา",    "color": "#fbbf24"},
    {"id": "negotiate", "label": "Negotiating",    "thai": "ต่อรองราคา",       "color": "#f97316"},
    {"id": "won",       "label": "Closed Won",     "thai": "ปิดการขาย",        "color": "#10b981"},
    {"id": "lost",      "label": "Closed Lost",    "thai": "เสียโอกาส",        "color": "#ef4444"},
    {"id": "hold",      "label": "On Hold",        "thai": "รอลูกค้า",         "color": "#a78bfa"},
]

TEAMS = [
    {"id": "all",     "name": "ALL TEAMS",    "thai": "รวมทุกทีม",   "color": "#0a1628"},
    {"id": "bp",      "name": "TEAM BP",      "thai": "ทีม BP",      "color": "#d97706"},
    {"id": "mai",     "name": "TEAM MAI",     "thai": "ทีม MAI",     "color": "#0891b2"},
    {"id": "oversea", "name": "TEAM OVERSEA", "thai": "ทีม Oversea", "color": "#7c3aed"},
]

STAGE_PRIORITY = {s["id"]: i for i, s in enumerate(STAGES)}

# =====================================================================
# SALESPERSON NAME RULES
# =====================================================================

# Names to exclude entirely (case-insensitive match on first word or full name)
SP_EXCLUDE = {"pond"}

# Canonical name mapping: any variant -> canonical name (team-aware key: "team:variant")
SP_NORMALIZE = {
    "bp:ping":   "PING",
    "bp:pilng":  "PING",
    "bp:piilng": "PING",
    "bp:piing":  "PING",
    "bp:piiing": "PING",   # covers 'PiiING'
    "mai:pun":   "PAN",
    "mai:pan":   "PAN",
}


def normalize_sp_name(team_id, raw_name):
    """
    Returns (canonical_name, excluded).
    excluded=True means this row should be skipped entirely.
    """
    name = str(raw_name).strip()
    first_word = name.split()[0].lower() if name else ""
    if first_word in SP_EXCLUDE or name.lower() in SP_EXCLUDE:
        return name, True
    key = f"{team_id}:{name.lower()}"
    if key in SP_NORMALIZE:
        return SP_NORMALIZE[key], False
    return name, False


# =====================================================================
# DATE HELPERS
# =====================================================================

def excel_serial_to_date(serial):
    if isinstance(serial, datetime):
        return serial.date()
    if isinstance(serial, (int, float)) and not math.isnan(serial):
        epoch = datetime(1899, 12, 30)
        return (epoch + timedelta(days=int(serial))).date()
    return None


def parse_date(val):
    """Parse any Excel date representation.  Rejects dates > 14 days future
    (avoids spurious future week labels from planned-shipment rows)."""
    if val is None:
        return None
    if isinstance(val, datetime):
        d = val.date()
    elif isinstance(val, (int, float)):
        try:
            d = excel_serial_to_date(val)
        except Exception:
            return None
    else:
        d = None

    if d is not None:
        if d > TODAY + timedelta(days=14) or d.year < 2020:
            return None
        return d

    if isinstance(val, str):
        val = val.strip()
        try:
            return datetime.fromisoformat(val).date()
        except Exception:
            pass
        # "MAY 1 26", "APR 27-2026", "MAY 1-2026", etc.
        m = re.match(r"([A-Za-z]+)[\s\-]+(\d{1,2})[\s\-]+(\d{2,4})$", val)
        if m:
            mon, day, year_s = m.groups()
            year = int(year_s) + 2000 if int(year_s) < 100 else int(year_s)
            for fmt in ["%b %d %Y", "%B %d %Y"]:
                try:
                    return datetime.strptime(f"{mon} {day} {year}", fmt).date()
                except Exception:
                    pass
        for fmt in ["%d/%m/%Y", "%m/%d/%Y", "%d-%m-%Y", "%Y/%m/%d"]:
            try:
                return datetime.strptime(val, fmt).date()
            except Exception:
                pass
    return None


def date_to_week_label(d):
    """Return ISO week label e.g. 'W08'."""
    if d is None:
        return None
    return f"W{d.isocalendar()[1]:02d}"


# =====================================================================
# STAGE / ACTIVITY MAPPING
# =====================================================================

def map_bp_mai_stage(status_val):
    if not status_val:
        return "lead"
    s = str(status_val).strip().lower()
    if "confirm" in s:
        return "won"
    if "waiting" in s or "decision" in s:
        return "negotiate"
    if "quotation" in s and "sent" in s:
        return "quote"
    if "visit" in s:
        return "visit"
    if "contact" in s:
        return "contact"
    return "lead"


def map_bp_mai_activity_bucket(stage):
    return {
        "won":       "newClients",
        "quote":     "quotations",
        "visit":     "visits",
        "contact":   "contacts",
        "negotiate": "contacts",
        "lead":      "contacts",
    }.get(stage, "contacts")


def map_ovs_stage(status_val):
    if not status_val:
        return "negotiate"
    s = str(status_val).strip().upper()
    if "WON" in s:
        return "won"
    if "LOST" in s:
        return "lost"
    return "negotiate"


def is_existing(val):
    if val is None:
        return False
    return "exist" in str(val).lower()


# =====================================================================
# WEEK ACTIVITY STRUCTURE
# =====================================================================

def empty_week():
    def sp():
        return {"existing": 0, "new": 0, "total": 0}
    return {
        "contacts":     sp(), "visits":      sp(), "quotations":   sp(),
        "problems":     sp(), "newClients":  sp(),
        "potentialTeu": sp(), "wonTeu":      sp(),
    }


def add_split(bucket, existing, count=1):
    if existing:
        bucket["existing"] += count
    else:
        bucket["new"] += count
    bucket["total"] += count


def merge_week(dst, src):
    """Add src week dict values into dst week dict (in-place)."""
    for field, val in src.items():
        if field == "week":
            continue
        if isinstance(val, dict):
            for sub in ("existing", "new", "total"):
                dst[field][sub] += val.get(sub, 0)


# =====================================================================
# SHAREPOINT DOWNLOAD  (Microsoft Graph + MSAL)
# =====================================================================

def _get_graph_token():
    """
    Return a Graph API access token.
    First run: device-code flow (user authenticates once via browser).
    Later runs: silent refresh from cached token.
    """
    try:
        import msal
    except ImportError:
        print("  WARN 'msal' not installed — run: pip install msal requests")
        return None

    os.makedirs(CACHE_DIR, exist_ok=True)
    cache = msal.SerializableTokenCache()
    if os.path.exists(TOKEN_CACHE):
        with open(TOKEN_CACHE, "r", encoding="utf-8") as fh:
            cache.deserialize(fh.read())

    app = msal.PublicClientApplication(
        MSAL_CLIENT_ID,
        authority=f"https://login.microsoftonline.com/{MSAL_TENANT}",
        token_cache=cache,
    )

    result = None
    accounts = app.get_accounts()
    if accounts:
        result = app.acquire_token_silent(MSAL_SCOPES, account=accounts[0])

    if not result or "access_token" not in result:
        print("\n  [MICROSOFT 365 LOGIN REQUIRED — first time only]")
        flow = app.initiate_device_flow(scopes=MSAL_SCOPES)
        if "user_code" not in flow:
            print(f"  ERROR Could not start device flow: {flow}")
            return None
        print(f"  1. Open this URL:  https://microsoft.com/devicelogin")
        print(f"  2. Enter the code: {flow['user_code']}")
        print(f"  Waiting for sign-in…  (tokens will be cached for ~90 days)")
        result = app.acquire_token_by_device_flow(flow)

    if "access_token" not in result:
        print(f"  WARN Auth failed: {result.get('error_description', 'unknown')}")
        return None

    if cache.has_state_changed:
        with open(TOKEN_CACHE, "w", encoding="utf-8") as fh:
            fh.write(cache.serialize())

    return result["access_token"]


def _graph_download(sharepoint_url, local_path):
    """
    Download a SharePoint file via Microsoft Graph API.
    Expected URL pattern:
      https://<tenant>.sharepoint.com/sites/<site>/Shared Documents/<folder>/<file>
    Returns True on success.
    """
    try:
        import requests as rq
    except ImportError:
        print("  WARN 'requests' not installed — run: pip install msal requests")
        return False

    token = _get_graph_token()
    if not token:
        return False

    m = re.match(
        r"https://([^/]+)(/sites/[^/]+)/Shared Documents(/.+)$",
        sharepoint_url,
    )
    if not m:
        print(f"  WARN Cannot parse SharePoint URL: {sharepoint_url}")
        return False

    from urllib.parse import quote
    hostname, site_path, file_path = m.groups()
    headers = {"Authorization": f"Bearer {token}"}

    # Step 1 — resolve site to get its stable Graph ID
    site_resp = rq.get(
        f"https://graph.microsoft.com/v1.0/sites/{hostname}:{site_path}",
        headers=headers, timeout=30,
    )
    if site_resp.status_code != 200:
        print(f"    WARN Site lookup {site_resp.status_code}: {site_resp.text[:150]}")
        return False
    site_id = site_resp.json().get("id", "")

    # Step 2 — download file using stable site ID + path
    file_enc = quote(file_path.lstrip("/"), safe="/")
    api_url = (
        f"https://graph.microsoft.com/v1.0"
        f"/sites/{site_id}/drive/root:/{file_enc}:/content"
    )
    resp = rq.get(api_url, headers=headers, allow_redirects=True, timeout=60)

    if resp.status_code == 200:
        os.makedirs(os.path.dirname(local_path), exist_ok=True)
        with open(local_path, "wb") as fh:
            fh.write(resp.content)
        print(f"    OK Downloaded {os.path.basename(local_path)}")
        return True
    else:
        print(f"    WARN Graph API {resp.status_code}: {resp.text[:200]}")
        return False


def ensure_file(sp_url, local_path, label=""):
    """
    Download file from SharePoint.  If download fails, use cached copy.
    Returns the local path if a usable file exists, else None.
    """
    fname = label or os.path.basename(local_path)
    print(f"  Fetching {fname} …")
    if _graph_download(sp_url, local_path):
        return local_path
    if os.path.isfile(local_path):
        print(f"  WARN Download failed — using cached {fname}")
        return local_path
    print(f"  ERROR {fname} unavailable — skipping this team")
    return None


# =====================================================================
# TEAM BP / MAI  — generic reader for the BP-format spreadsheet
# =====================================================================

def read_team_file(filepath, team_id, preferred_sheets,
                   sps_map, wk_data, cust_map, sp_seq, events):
    """
    Read one team's BP-format Excel file and accumulate results into the
    caller's shared dicts (sps_map, wk_data, cust_map) and events list.

    preferred_sheets : list of sheet names to try, in order.
                       Falls back to the first sheet that looks like data.
    """
    print(f"  Reading {os.path.basename(filepath)} ({team_id.upper()}) …")

    wb = openpyxl.load_workbook(filepath, data_only=True, read_only=True)

    # Pick the right sheet
    sheet_name = None
    for candidate in (preferred_sheets or []):
        if candidate in wb.sheetnames:
            sheet_name = candidate
            break
    if sheet_name is None:
        for sn in wb.sheetnames:
            sl = sn.lower()
            if re.match(r"202\d", sn) or "sale" in sl or "log" in sl:
                sheet_name = sn
                break
    if sheet_name is None:
        sheet_name = wb.sheetnames[0]
    print(f"    Sheet: '{sheet_name}'")

    ws = wb[sheet_name]
    row_count = 0

    for row in ws.iter_rows(min_row=2, values_only=True):
        # Column layout (0-indexed):
        # A=week  B=input_date  C=sales_name  D=customer  E=cust_status
        # F=job_type  G=grp_comm  H=commodity
        # I=country_orig  J=pol  K=country_dest  L=pod
        # M=type  N=size  O=potential_teu  P=term  Q=status  R=details
        if len(row) < 17:
            continue
        date_raw          = row[1]
        sales_name        = row[2]
        customer          = row[3]
        cust_status       = row[4]
        grp_comm          = row[6]
        commodity         = row[7]
        pod               = row[11]
        potential_teu_raw = row[14]
        status_raw        = row[16]
        details           = row[17] if len(row) > 17 else None

        if not sales_name:
            continue
        sales_name = str(sales_name).strip()
        if not sales_name or sales_name.lower() == "sales name":
            continue
        sales_name, excluded = normalize_sp_name(team_id, sales_name)
        if excluded:
            continue

        # Register salesperson
        sp_key = f"{team_id}_{sales_name.upper()}"
        if sp_key not in sps_map:
            sp_seq[0] += 1
            sp_id = f"s{sp_seq[0]}"
            initials = "".join(p[0].upper() for p in sales_name.split()[:2])
            sps_map[sp_key] = {
                "id": sp_id, "name": sales_name, "thai": sales_name,
                "team": team_id, "avatar": initials[:2],
            }
            wk_data[sp_id] = {}
        sp_id = sps_map[sp_key]["id"]

        # Week label
        d = parse_date(date_raw)
        if not d:
            continue
        wk_label = date_to_week_label(d)

        if wk_label not in wk_data[sp_id]:
            wk_data[sp_id][wk_label] = empty_week()
        wk = wk_data[sp_id][wk_label]

        existing = is_existing(cust_status)
        stage    = map_bp_mai_stage(status_raw)
        bucket   = map_bp_mai_activity_bucket(stage)
        add_split(wk[bucket], existing)

        teu = 0
        if potential_teu_raw and isinstance(potential_teu_raw, (int, float)):
            if not math.isnan(potential_teu_raw):
                teu = int(potential_teu_raw)
        if stage == "won":
            add_split(wk["wonTeu"], existing, teu)
        else:
            add_split(wk["potentialTeu"], existing, teu)

        # Row-level event (for drill-down on the dashboard)
        cust_for_event = str(customer).strip() if customer else "(no customer)"
        events.append({
            "team":     team_id,
            "spId":     sp_id,
            "customer": cust_for_event,
            "week":     wk_label,
            "date":     d.isoformat(),
            "type":     bucket,
            "stage":    stage,
            "teu":      teu,
            "existing": existing,
            "notes":    (str(details).strip() if details else "")[:140],
        })

        # Customer record
        if customer:
            cust_str = str(customer).strip()
            if not cust_str:
                continue
            cust_key = f"{team_id}_{cust_str}"
            location = str(pod).strip() if pod else ""
            industry = (str(grp_comm).strip() if grp_comm
                        else (str(commodity).strip() if commodity else ""))
            date_str = d.isoformat()
            note     = str(details).strip() if details else ""

            if cust_key not in cust_map:
                cust_map[cust_key] = {
                    "name": cust_str, "team": team_id,
                    "stage": stage, "location": location,
                    "industry": industry, "lastActivity": date_str,
                    "firstContactDate": date_str, "owner": sp_id,
                    "potentialTeu": teu, "notes": note,
                    "sinceWeek": wk_label,
                    "contactsThisMonth": 0, "quotationsSent": 0,
                    "_acts": [bucket],
                }
            else:
                c = cust_map[cust_key]
                if STAGE_PRIORITY.get(stage, 0) > STAGE_PRIORITY.get(c["stage"], 0):
                    c["stage"] = stage
                if date_str and date_str > c["lastActivity"]:
                    c["lastActivity"] = date_str
                    c["owner"] = sp_id
                if date_str and (not c["firstContactDate"]
                                 or date_str < c["firstContactDate"]):
                    c["firstContactDate"] = date_str
                    c["sinceWeek"] = wk_label
                c["potentialTeu"] = max(c["potentialTeu"], teu)
                if note and not c["notes"]:
                    c["notes"] = note
                c["_acts"].append(bucket)
        row_count += 1

    wb.close()
    print(f"    OK {row_count} activity rows")


def read_bp_mai(bp_path, mai_path):
    """
    Read Team BP (sheet '2026') and Team MAI ('Sales Log' file).
    Returns (sps_list, wk_data_dict, customers_list, events_list).
    """
    sps_map  = {}
    wk_data  = {}
    cust_map = {}
    sp_seq   = [0]
    events   = []

    if bp_path:
        read_team_file(bp_path, "bp", ["2026"],
                       sps_map, wk_data, cust_map, sp_seq, events)
    else:
        print("  WARN Team BP file unavailable — skipping")

    if mai_path:
        yr = str(datetime.today().year)
        read_team_file(mai_path, "mai",
                       [f"Sales Log {yr}", yr, "Sales Log 2026", "Sales Log", "MAI"],
                       sps_map, wk_data, cust_map, sp_seq, events)
    else:
        print("  WARN Team MAI file unavailable — skipping")

    custs = []
    for c in cust_map.values():
        acts = c.pop("_acts", [])
        c["contactsThisMonth"] = acts.count("contacts") + acts.count("visits")
        c["quotationsSent"]    = acts.count("quotations")
        custs.append(c)

    return list(sps_map.values()), wk_data, custs, events


# =====================================================================
# TEAM OVERSEA  — scans ALL monthly Excel files in OVS_BASE folder
# =====================================================================

def _build_col_map(ws):
    """Read rows 2-3, return {HEADER_TEXT: col_index (0-based)}."""
    col_map = {}
    try:
        row2 = [c.value for c in ws[2]]
        row3 = [c.value for c in ws[3]]
    except Exception:
        return col_map
    for i, val in enumerate(row3):
        if val:
            col_map[str(val).strip().upper()] = i
    for i, val in enumerate(row2):
        if val:
            k = str(val).strip().upper()
            if k not in col_map:
                col_map[k] = i
    return col_map


def _ci(col_map, *keywords):
    """Return first column index whose key contains any of the keywords."""
    for kw in keywords:
        kw_up = kw.upper()
        for k, v in col_map.items():
            if kw_up in k:
                return v
    return None


def _safe_int(row, idx):
    if idx is None or idx >= len(row):
        return 0
    v = row[idx]
    if isinstance(v, (int, float)) and not math.isnan(v):
        return int(v)
    return 0


def find_oversea_files(base_dir):
    """Recursively find all .xlsx files under base_dir (skips ~$ temp files)."""
    files = []
    if not os.path.isdir(base_dir):
        print(f"  WARN Oversea base dir not found:\n    {base_dir}")
        return files
    for root, dirs, filenames in os.walk(base_dir):
        dirs[:] = sorted(d for d in dirs
                         if not d.startswith(".") and d != "desktop.ini")
        for fn in sorted(filenames):
            if fn.endswith(".xlsx") and not fn.startswith("~$"):
                files.append(os.path.join(root, fn))
    return files


def _read_oversea_file(filepath, sps_registry, wk_accum, cust_accum, sp_seq, events):
    """
    Process one Oversea xlsx file.  Accumulates into shared dicts:
      sps_registry : {sp_name → sp_dict}
      wk_accum     : {sp_name → {wk_label → week_dict}}
      cust_accum   : {"oversea_<name>" → customer_dict}
      events       : flat list of row-level activity events
    """
    rel = os.path.join(
        os.path.basename(os.path.dirname(filepath)),
        os.path.basename(filepath),
    )
    print(f"  Reading {rel} …")

    wb = openpyxl.load_workbook(filepath, data_only=True, read_only=False)
    sales_sheets = [s for s in wb.sheetnames
                    if re.match(r"sales\s+[a-h]", s.lower())]

    if not sales_sheets:
        print(f"    WARN No 'Sales X' sheets found — skipping")
        wb.close()
        return

    row_total = 0
    for sheet_name in sales_sheets:
        ws = wb[sheet_name]

        # Salesperson name from "Sales A - NIDHI" → "Nidhi"
        m = re.search(r"[-–]\s*(.+)$", sheet_name)
        sp_name = m.group(1).strip().title() if m else sheet_name.strip().title()

        # Register SP (deduplicate across files by name)
        if sp_name not in sps_registry:
            sp_seq[0] += 1
            sp_id = f"s{sp_seq[0]}"
            sps_registry[sp_name] = {
                "id": sp_id, "name": sp_name, "thai": sp_name,
                "team": "oversea", "avatar": sp_name[:2].upper(),
            }
            wk_accum[sp_name] = {}
        sp_id = sps_registry[sp_name]["id"]

        col_map   = _build_col_map(ws)
        idx_no    = _ci(col_map, "NO.")
        idx_date  = _ci(col_map, "ENTRY DATE")
        idx_ctype = _ci(col_map, "CUSTOMER TYPE")
        idx_20gp  = _ci(col_map, "20GP")
        idx_40hq  = _ci(col_map, "40HQ")
        idx_20rh  = _ci(col_map, "20RH")
        idx_40rh  = _ci(col_map, "40RH")
        idx_lcust = _ci(col_map, "LOCAL CUSTOMER")
        idx_agent = _ci(col_map, "OVERSEAS AGENT NAME", "AGENT NAME")
        idx_actry = _ci(col_map, "AGENT COUNTRY")
        idx_stat  = _ci(col_map, "STATUS")
        idx_rem   = _ci(col_map, "REMARKS")
        idx_pod   = _ci(col_map, "POD")

        for row in ws.iter_rows(min_row=5, values_only=True):
            # Skip blank / footer rows
            if idx_no is not None and idx_no < len(row):
                no_val = row[idx_no]
                if no_val is None:
                    continue
                if isinstance(no_val, str) and re.search(
                        r"total|count|margin", no_val, re.I):
                    continue

            date_raw = row[idx_date] if idx_date is not None and idx_date < len(row) else None
            d = parse_date(date_raw)
            if not d:
                continue
            wk_label  = date_to_week_label(d)
            ctype_raw = row[idx_ctype] if idx_ctype is not None and idx_ctype < len(row) else None
            existing  = is_existing(ctype_raw)

            local_cust = row[idx_lcust] if idx_lcust is not None and idx_lcust < len(row) else None
            agent_name = row[idx_agent] if idx_agent is not None and idx_agent < len(row) else None
            agent_ctry = row[idx_actry] if idx_actry is not None and idx_actry < len(row) else None

            is_agent = bool(agent_name and str(agent_name).strip())
            if is_agent:
                cust_name = str(agent_name).strip() + " [Agent]"
                location  = str(agent_ctry).strip() if agent_ctry else "International"
                industry  = "Overseas Agent"
            elif local_cust and str(local_cust).strip():
                cust_name = str(local_cust).strip()
                pod_raw   = row[idx_pod] if idx_pod is not None and idx_pod < len(row) else None
                location  = str(pod_raw).strip() if pod_raw else "Thailand"
                industry  = "Local Customer"
            else:
                continue

            # TEU: 20GP×1 + 40HQ×2 + 20RH×1 + 40RH×2
            teu = (_safe_int(row, idx_20gp)
                   + _safe_int(row, idx_40hq) * 2
                   + _safe_int(row, idx_20rh)
                   + _safe_int(row, idx_40rh) * 2)

            status_raw = row[idx_stat] if idx_stat is not None and idx_stat < len(row) else None
            stage      = map_ovs_stage(status_raw)
            remark_raw = row[idx_rem]  if idx_rem  is not None and idx_rem  < len(row) else None
            notes      = str(remark_raw).strip() if remark_raw else ""

            # Week activity
            if wk_label not in wk_accum[sp_name]:
                wk_accum[sp_name][wk_label] = empty_week()
            wk = wk_accum[sp_name][wk_label]

            add_split(wk["contacts"], existing)
            if stage == "won":
                add_split(wk["newClients"], existing)
                add_split(wk["wonTeu"], existing, teu)
            elif stage != "lost":
                add_split(wk["potentialTeu"], existing, teu)

            # Row-level event (Oversea always counts as 'contacts'; if won → also wonDeals)
            ev_type = "newClients" if stage == "won" else "contacts"
            events.append({
                "team":     "oversea",
                "spId":     sp_id,
                "customer": cust_name,
                "week":     wk_label,
                "date":     d.isoformat(),
                "type":     ev_type,
                "stage":    stage,
                "teu":      teu,
                "existing": existing,
                "notes":    (notes or "")[:140],
            })

            # Customer record
            cust_key = f"oversea_{cust_name}"
            date_str = d.isoformat()
            if cust_key not in cust_accum:
                cust_accum[cust_key] = {
                    "name": cust_name, "team": "oversea",
                    "stage": stage, "location": location,
                    "industry": industry, "lastActivity": date_str,
                    "firstContactDate": date_str, "owner": sp_id,
                    "potentialTeu": teu, "notes": notes,
                    "sinceWeek": wk_label,
                    "contactsThisMonth": 1, "quotationsSent": 0,
                }
            else:
                c = cust_accum[cust_key]
                if STAGE_PRIORITY.get(stage, 0) > STAGE_PRIORITY.get(c["stage"], 0):
                    c["stage"] = stage
                if date_str > c["lastActivity"]:
                    c["lastActivity"] = date_str
                    c["owner"] = sp_id
                if date_str < c["firstContactDate"]:
                    c["firstContactDate"] = date_str
                    c["sinceWeek"] = wk_label
                c["potentialTeu"] = max(c["potentialTeu"], teu)
                if notes and not c["notes"]:
                    c["notes"] = notes
                c["contactsThisMonth"] += 1
            row_total += 1

    wb.close()
    print(f"    OK {os.path.basename(filepath)}: {row_total} deal rows "
          f"across {len(sales_sheets)} sheet(s)")


def read_all_oversea(base_dir, sp_id_start):
    """
    Scan base_dir recursively for monthly xlsx files.
    Deduplicates salespeople by name across files.
    Returns (sps_list, wk_data_by_sp_id, customers_list).
    """
    files = find_oversea_files(base_dir)
    if not files:
        print(f"  WARN No Oversea xlsx files found in:\n    {base_dir}")
        return [], {}, [], []

    print(f"\n--- Team OVERSEA ({len(files)} file(s)) ---")

    sps_registry = {}   # sp_name → sp_dict
    wk_accum     = {}   # sp_name → {wk_label → week_dict}
    cust_accum   = {}   # "oversea_<name>" → customer_dict
    events       = []
    sp_seq       = [sp_id_start]

    for filepath in files:
        try:
            _read_oversea_file(filepath, sps_registry, wk_accum, cust_accum, sp_seq, events)
        except Exception as exc:
            print(f"    WARN {os.path.basename(filepath)}: {exc}")

    # Convert wk_accum (keyed by sp_name) → keyed by sp_id
    sps     = list(sps_registry.values())
    wk_data = {sp["id"]: wk_accum.get(sp["name"], {}) for sp in sps}
    custs   = list(cust_accum.values())
    return sps, wk_data, custs, events


# =====================================================================
# MERGE & FINALISE
# =====================================================================

def finalise_customers(raw_customers):
    today = TODAY
    seen  = set()
    out   = []
    for c in raw_customers:
        name = c.get("name", "").strip()
        if not name or len(name) < 2:
            continue
        key = f"{c.get('team', '')}_{name}"
        if key in seen:
            continue
        seen.add(key)
        try:
            fc = datetime.fromisoformat(c["firstContactDate"]).date()
        except Exception:
            fc = today
        c["daysSinceFirstContact"] = max(0, (today - fc).days)
        c["id"] = f"C{len(out)+1:03d}"
        out.append(c)
    return out


def build_all_weeks(all_salespeople, bp_mai_wk, ovs_wk):
    """Collect all week labels, trim to last 12, fill gaps with zeros."""
    combined = {}
    for sp in all_salespeople:
        sid  = sp["id"]
        src  = bp_mai_wk if sp["team"] in ("bp", "mai") else ovs_wk
        combined[sid] = src.get(sid, {})

    all_labels = set()
    for wks in combined.values():
        all_labels.update(wks.keys())

    def wk_num(w):
        m = re.match(r"W(\d+)", w)
        return int(m.group(1)) if m else 0

    sorted_labels = sorted(all_labels, key=wk_num)
    if len(sorted_labels) > 12:
        sorted_labels = sorted_labels[-12:]

    activity = {}
    for sp in all_salespeople:
        sid   = sp["id"]
        sp_wk = combined.get(sid, {})
        rows  = []
        for wl in sorted_labels:
            entry = sp_wk.get(wl, empty_week())
            entry["week"] = wl
            rows.append(entry)
        activity[sid] = rows

    return sorted_labels, activity


# =====================================================================
# MARKET INTELLIGENCE  — reads local market_intelligence.xlsx
# =====================================================================

_PRODUCT_CATS = [
    ("chicken", ["chicken", "poultry", "duck", "broiler", "layer", "hen"]),
    ("seafood",  ["seafood", "fish", "tuna", "shrimp", "prawn", "crab",
                  "squid", "sardine", "lobster", "oyster", "anchovy", "mackerel"]),
    ("rice",    ["rice", "jasmine", "glutinous", "hom mali", "parboiled"]),
    ("sugar",   ["sugar", "molasses", "syrup", "sucrose"]),
]


def _cat_from_products(products):
    cats = set()
    for p in products:
        pl = p.lower()
        for cat_id, kws in _PRODUCT_CATS:
            if any(k in pl for k in kws):
                cats.add(cat_id)
                break
    if len(cats) == 1:
        return cats.pop()
    if len(cats) > 1:
        return "multi"
    return "multi"


def read_market_intelligence(path):
    """
    Parse market_intelligence.xlsx  (columns: Shipper / Product / Country / Shipments).
    Returns a MARKET dict ready to embed in data.json, or None if file missing.
    """
    if not os.path.isfile(path):
        print(f"  WARN Market Intelligence file not found: {path}")
        return None

    print(f"  Reading {os.path.basename(path)} (Market Intelligence) …")
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.worksheets[0]

    # {shipper_name: {cv: {country: shipments}, products: set}}
    shipper_map = {}
    country_totals = {}
    product_set = set()
    row_count = 0

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue
        name     = str(row[0]).strip()
        product  = str(row[1]).strip() if len(row) > 1 and row[1] else None
        country  = str(row[2]).strip() if len(row) > 2 and row[2] else None
        ships    = row[3] if len(row) > 3 else 0
        try:
            ships = int(ships) if ships and not (isinstance(ships, float) and math.isnan(ships)) else 0
        except (TypeError, ValueError):
            ships = 0

        if not name or not product or not country:
            continue

        if name not in shipper_map:
            shipper_map[name] = {"cv": {}, "products": set()}
        s = shipper_map[name]
        s["cv"][country]    = s["cv"].get(country, 0) + ships
        s["products"].add(product)
        country_totals[country] = country_totals.get(country, 0) + ships
        product_set.add(product)
        row_count += 1

    wb.close()
    print(f"    OK {row_count} rows -> {len(shipper_map)} unique shippers")

    shippers = []
    for i, (name, data) in enumerate(
            sorted(shipper_map.items(), key=lambda x: -sum(x[1]["cv"].values()))):
        prods   = sorted(data["products"])
        cv      = data["cv"]
        total   = sum(cv.values())
        countries_sorted = sorted(cv.keys(), key=lambda c: -cv[c])
        cat     = _cat_from_products(prods)
        shippers.append({
            "id":       f"m{i+1:03d}",
            "name":     name,
            "nameTh":   None,
            "cat":      cat,
            "products": prods,
            "cv":       cv,
            "countries": countries_sorted,
            "totalTeu": total,
        })

    countries = [
        {"country": c, "shipments": v}
        for c, v in sorted(country_totals.items(), key=lambda x: -x[1])
    ]

    return {
        "SHIPPERS":  shippers,
        "COUNTRIES": countries,
        "PRODUCTS":  sorted(product_set),
        "isMock":    False,
    }


# =====================================================================
# STATUS FILE  (used by --fallback mode)
# =====================================================================

def _mark_run(status):
    os.makedirs(CACHE_DIR, exist_ok=True)
    with open(STATUS_FILE, "w", encoding="utf-8") as fh:
        fh.write(f"{status} {datetime.now().isoformat()}\n")


def _last_run_ok():
    """Return True if the last run was successful and happened within 48 h."""
    if not os.path.isfile(STATUS_FILE):
        return False
    with open(STATUS_FILE, "r", encoding="utf-8") as fh:
        line = fh.readline().strip()
    if not line.startswith("SUCCESS"):
        return False
    try:
        ts = datetime.fromisoformat(line.split()[1])
        return (datetime.now() - ts).total_seconds() < 48 * 3600
    except Exception:
        return False


# =====================================================================
# MAIN
# =====================================================================

def _market_only():
    """Update just the MARKET section of an existing data.json without touching SharePoint."""
    print("\n=== Market Intelligence update (--market-only) ===\n")
    market = read_market_intelligence(MARKET_PATH)
    if not market:
        print("ERROR Market file not found — nothing written.")
        sys.exit(1)

    existing = {}
    if os.path.isfile(OUTPUT_PATH):
        with open(OUTPUT_PATH, "r", encoding="utf-8") as fh:
            existing = json.load(fh)

    existing["MARKET"] = market
    existing["generated_at"] = datetime.utcnow().isoformat() + "Z"

    out_dir = os.path.dirname(OUTPUT_PATH)
    if out_dir:
        os.makedirs(out_dir, exist_ok=True)
    with open(OUTPUT_PATH, "w", encoding="utf-8") as fh:
        json.dump(existing, fh, ensure_ascii=False, indent=2, default=str)

    print(f"\nDONE  MARKET section updated  ->  {OUTPUT_PATH}")
    print(f"      Shippers : {len(market['SHIPPERS'])}")
    print(f"      Countries: {len(market['COUNTRIES'])}")
    print(f"      Products : {market['PRODUCTS']}")

    if AUTO_GIT_PUSH:
        print("\nPushing to GitHub…")
        cmds = [
            ["git", "-C", _REPO_ROOT, "add", "data.json"],
            ["git", "-C", _REPO_ROOT, "commit", "-m",
             f"data: update market intelligence {datetime.utcnow().strftime('%Y-%m-%d')}"],
            ["git", "-C", _REPO_ROOT, "push"],
        ]
        for cmd in cmds:
            r = subprocess.run(cmd, capture_output=True, text=True)
            if r.returncode != 0:
                print(f"  WARN git: {r.stderr.strip()}")
            else:
                print(f"  OK  {' '.join(cmd[2:])}")
    print("\nDone.\n")


def main():
    if "--market-only" in sys.argv:
        _market_only()
        return

    # --fallback mode: skip if last sync was already successful this week
    if "--fallback" in sys.argv:
        if _last_run_ok():
            print("Fallback check: last sync was recent — nothing to do.")
            return
        print("Fallback mode: last sync was not successful — running now.\n")

    print("\n=== Sales Activity Console — Excel Sync ===\n")
    _mark_run("RUNNING")

    try:
        # ── Step 1 : Download BP / MAI from SharePoint ────────────────────
        print("--- Team BP & MAI ---")
        bp_path  = ensure_file(BP_SHAREPOINT_URL,  BP_PATH,  "Team BP")
        mai_path = ensure_file(MAI_SHAREPOINT_URL, MAI_PATH, "Team MAI")

        bp_mai_sps, bp_mai_wk, bp_mai_custs, bp_mai_events = read_bp_mai(bp_path, mai_path)
        max_seq = max((int(sp["id"][1:]) for sp in bp_mai_sps), default=0)

        # ── Step 2 : Read all Oversea monthly files ───────────────────────
        ovs_sps, ovs_wk, ovs_custs, ovs_events = read_all_oversea(OVS_BASE, max_seq)

        # ── Step 2b : Market Intelligence ─────────────────────────────────
        print("\n--- Market Intelligence ---")
        market_data = read_market_intelligence(MARKET_PATH)

        # ── Step 3 : Merge & write ────────────────────────────────────────
        all_sps   = bp_mai_sps + ovs_sps
        all_custs = finalise_customers(bp_mai_custs + ovs_custs)

        sorted_weeks, activity = build_all_weeks(all_sps, bp_mai_wk, ovs_wk)

        # Events: keep only those that fall in the displayed weeks (matches WEEKS array).
        weeks_set = set(sorted_weeks)
        all_events = [
            e for e in (bp_mai_events + ovs_events)
            if e.get("week") in weeks_set
        ]
        # Sort newest first for nicer drill-down lists
        all_events.sort(key=lambda e: (e.get("date") or "", e.get("week") or ""), reverse=True)

        data = {
            "TEAMS":       TEAMS,
            "SALESPEOPLE": all_sps,
            "WEEKS":       sorted_weeks,
            "ACTIVITY":    activity,
            "STAGES":      STAGES,
            "CUSTOMERS":   all_custs,
            "EVENTS":      all_events,
            "MARKET":      market_data,
            "generated_at": datetime.utcnow().isoformat() + "Z",
        }

        out_dir = os.path.dirname(OUTPUT_PATH)
        if out_dir:
            os.makedirs(out_dir, exist_ok=True)
        with open(OUTPUT_PATH, "w", encoding="utf-8") as fh:
            json.dump(data, fh, ensure_ascii=False, indent=2, default=str)

        print(f"\nDONE  data.json written  ->  {OUTPUT_PATH}")
        print(f"      Salespeople : {len(all_sps)}")
        print(f"      Customers   : {len(all_custs)}")
        print(f"      Events      : {len(all_events)}")
        print(f"      Weeks       : {sorted_weeks}")
        print(f"      Generated   : {data['generated_at']}")

        # ── Optional: Git push ────────────────────────────────────────────
        if AUTO_GIT_PUSH:
            print("\nPushing to GitHub…")
            cmds = [
                ["git", "-C", _REPO_ROOT, "add", "data.json"],
                ["git", "-C", _REPO_ROOT, "commit", "-m",
                 f"data: auto-sync {datetime.utcnow().strftime('%Y-%m-%d %H:%M')} UTC"],
                ["git", "-C", _REPO_ROOT, "push"],
            ]
            for cmd in cmds:
                r = subprocess.run(cmd, capture_output=True, text=True)
                if r.returncode != 0:
                    print(f"  WARN git: {r.stderr.strip()}")
                else:
                    print(f"  OK  {' '.join(cmd[2:])}")
            print("DONE  GitHub push complete")

        _mark_run("SUCCESS")
        print("\nSync complete.\n")

    except Exception as exc:
        _mark_run("FAILED")
        print(f"\nERROR during sync: {exc}")
        raise


if __name__ == "__main__":
    main()
